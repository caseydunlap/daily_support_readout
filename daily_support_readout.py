#Import packages
import pandas as pd
import numpy as np
import pytz
import boto3
import json
import os
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
import matplotlib.pyplot as plt
import seaborn as sns
import requests
import snowflake.connector
from sqlalchemy import create_engine
from decimal import Decimal
from requests.auth import HTTPBasicAuth
from openpyxl.utils import get_column_letter
import urllib
import shutil
import smtplib
import openpyxl
from openpyxl.drawing.image import Image as XLSXImage
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import ssl
from PIL import Image, ImageDraw, ImageFont
import matplotlib.gridspec as gridspec

#Create Boto3 Session + Pass AWS Keys
with open('C:\\Users\\mdunlap\\Desktop\\Daily Support ReadOut\\Secrets\\aws_access_key.txt', 'r') as aws_access_key_temp:
    aws_access_key_id = aws_access_key_temp.read().replace('\n', '')
with open('C:\\Users\\mdunlap\\Desktop\\Daily Support ReadOut\\Secrets\\aws_secret_key.txt', 'r') as aws_secret_access_key_temp:
    aws_secret_access_key = aws_secret_access_key_temp.read().replace('\n', '')
region_name = 'us-east-1'

session = boto3.Session(
    aws_access_key_id=aws_access_key_id,
    aws_secret_access_key=aws_secret_access_key,
    region_name=region_name)

#Standardize formatting for plots
sns.set_style("white")
sns.set_context("talk")

#Fetch Queue Names from AWS - Store in dataframe
connect = session.client('connect',region_name='us-east-1')
with open('C:\\Users\\mdunlap\\Desktop\\Daily Support ReadOut\\Secrets\\aws_instance_id.txt', 'r') as instance_id_temp:
    instance_id = instance_id_temp.read().replace('\n', '')
queues = connect.list_queues(InstanceId=instance_id, QueueTypes=['STANDARD'])
queue_df = pd.DataFrame(queues['QueueSummaryList'])

#Fetch Agent Names, paginate if needed - Store in dataframe
connect_client = session.client('connect', endpoint_url='https://connect.us-east-1.amazonaws.com')
paginator = connect_client.get_paginator('list_users')
user_params = {'InstanceId': instance_id}
agent_list = []
for page in paginator.paginate(**user_params):
    agent_list.extend(page['UserSummaryList'])
agent_df = pd.DataFrame(agent_list, columns=['Id', 'Username'])

#Fetch Contacts Queued by Hour From Yesterday - Store in dataframe
connect_client = session.client('connect', endpoint_url='https://connect.us-east-1.amazonaws.com')
eastern = pytz.timezone('US/Eastern')
now = datetime.now(eastern)
yesterday_date = now - timedelta(days=1)
start_time = datetime(yesterday_date.year, yesterday_date.month, yesterday_date.day, 0, 0, 0, tzinfo=eastern)
end_time = datetime(yesterday_date.year, yesterday_date.month, yesterday_date.day, 23, 59, 59, tzinfo=eastern)
final_data = []
queue_response = connect_client.list_queues(InstanceId=instance_id, QueueTypes=['STANDARD'])
queue_ids = [queue['Id'] for queue in queue_response['QueueSummaryList']]
metrics = ['ContactCount', 'ContactDuration', 'ContactQueueTime']

with open('C:\\Users\\mdunlap\\Desktop\\Daily Support ReadOut\\Secrets\\aws_resource_arn.txt', 'r') as resource_arn_temp:
    resource_arn_2 = resource_arn_temp.read().replace('\n', '')

current_date = start_time
while current_date <= end_time:
    start_time_unix = int(current_date.timestamp())
    end_time_unix = int((current_date + timedelta(hours=1)).timestamp())

    agent_data = []

    next_token = None
    while True:
        if next_token:
            response = connect_client.get_metric_data_v2(
                ResourceArn=resource_arn_2,
                StartTime=start_time_unix,
                EndTime=end_time_unix,
                Filters=[
                    {
                        'FilterKey': 'QUEUE',
                        'FilterValues': queue_ids
                    },
                ],
                Groupings=[
                    'QUEUE'
                ],
                Metrics=[
                    {
                        'Name': 'CONTACTS_QUEUED',
                        'Threshold': [
                            {
                                'Comparison': 'LT',
                                'ThresholdValue': 100
                            },
                        ]
                    },
                ],
                NextToken=next_token,
                MaxResults=100
            )
        else:
            response = connect_client.get_metric_data_v2(
                ResourceArn=resource_arn_2,
                StartTime=start_time_unix,
                EndTime=end_time_unix,
                Filters=[
                    {
                        'FilterKey': 'QUEUE',
                        'FilterValues': queue_ids
                    },
                ],
                Groupings=[
                    'QUEUE','AGENT'
                ],
                Metrics=[
                    {
                        'Name': 'CONTACTS_QUEUED',
                        'Threshold': [
                            {
                                'Comparison': 'LT',
                                'ThresholdValue': 100
                            },
                        ]
                    },
                ],
                MaxResults=100
            )

            for metric in response['MetricResults']:
                queue_id = metric['Dimensions']['QUEUE']
                agent_id = metric['Dimensions']['AGENT']
                contacts_queued = metric['Collections'][0]['Value']
                final_data.append({'Queue ID': queue_id,'Agent ID':agent_id, 'Contacts Queued': contacts_queued, 'Timestamp': current_date})

            if 'NextToken' in response:
                next_token = response['NextToken']
            else:
                break

    current_date += timedelta(hours=1)

phone_df = pd.DataFrame(final_data)

#Append queue names and agent names to new dataframe, remove some unnessecary queues
result = pd.merge(phone_df, agent_df, left_on='Agent ID', right_on = 'Id',how='left')
result = pd.merge(result, queue_df, left_on='Queue ID', right_on = 'Id',how='left')
result.drop(['Id_x', 'Id_y', 'Arn','QueueType'], axis=1, inplace=True)
names_to_drop = ['testqueue', 'Offshore Support - Chat Only', 'Onshore Support Chat']
result = result[~result['Name'].isin(names_to_drop)]

#Aggregate result dataframe into new dataframe, grouped by queue name
grouped_phone_queue= result.groupby(['Name'])['Contacts Queued'].sum().reset_index()

#Create first plot image - Contacts queued by queue
grouped_phone_queue = grouped_phone_queue.sort_values(by='Contacts Queued', ascending=False)
grouped_phone_queue.set_index('Name', inplace=True)
grouped_phone_queue = grouped_phone_queue.astype(int)
plt.figure(figsize=(10, 5))
sns.heatmap(grouped_phone_queue, annot=True, fmt="d", cmap="coolwarm", cbar=False)
plt.title('Contacts Queued by Queue')
plt.xticks([])
plt.ylabel('')
plt.tight_layout()
plt.savefig('contacts_queued.png')

#Fetch Live Chat, Portal, Email, Ticket status data from JIRA
jira_url = "https://hhaxsupport.atlassian.net"
api_endpoint = "/rest/api/3/search"
jql_query = """project = '10000' AND created >= startOfDay(-1d) AND created < startOfDay()"""
jql_query_encoded = urllib.parse.quote(jql_query)
startAt = 0
maxResults = 100

with open('C:\\Users\\mdunlap\\Desktop\\Daily Support ReadOut\\Secrets\\jira_email.txt', 'r') as temp_jira_email:
    jira_email = temp_jira_email.read().replace('\n', '')

with open('C:\\Users\\mdunlap\\Desktop\\Daily Support ReadOut\\Secrets\\jira_apikey.txt', 'r') as temp_jira_apikey:
    jira_key = temp_jira_apikey.read().replace('\n', '')

all_issues = []

while True:
    api_url = f"{jira_url}{api_endpoint}?jql={jql_query_encoded}&startAt={startAt}&maxResults={maxResults}"

    response = requests.get(
        api_url,
        auth=HTTPBasicAuth(jira_email, jira_key),
        headers={
            "Accept": "application/json"
        }
    )

   
    json_response = response.json()

    
    print(f"Status code: {response.status_code}")

    
    if response.status_code == 200:
        all_issues.extend(json_response['issues'])
        
        if json_response['total'] == len(all_issues):
            break
        else:
            startAt += maxResults
    else:
        break

#Fetch data for aging backlog report
jira_url_backlog = "https://hhaxsupport.atlassian.net"
api_endpoint_backlog = "/rest/api/3/search"
maxResults_backlog = 100 

end_date_backlog = datetime.strptime("04-Sep-2023", "%d-%b-%Y")
start_date_backlog = datetime.strptime("18-Aug-2023", "%d-%b-%Y")
N = (end_date_backlog - start_date_backlog).days
dates = [(start_date_backlog + timedelta(days=i)).strftime('%d-%b-%Y') for i in range(N)]

backlog_df_temp = []

previous_started = 3286

for date in dates:
    current_day_start = datetime.strptime(date, '%d-%b-%Y').strftime('%Y-%m-%d')
    next_day_start = (datetime.strptime(date, '%d-%b-%Y') + timedelta(days=1)).strftime('%Y-%m-%d')
    created_issues = []
    resolved_issues = []
    startAt_backlog = 0

    while True:
        jql_query_backlog = f"project = '10000' AND created >= '{current_day_start}' AND created < '{next_day_start}'"
        jql_query_encoded_backlog = urllib.parse.quote(jql_query_backlog)
        api_url_backlog = f"{jira_url_backlog}{api_endpoint_backlog}?jql={jql_query_encoded_backlog}&startAt={startAt_backlog}&maxResults={maxResults_backlog}"

        backlog_response = requests.get(
            api_url_backlog,
            auth=HTTPBasicAuth(jira_email, jira_key),
            headers={"Accept": "application/json"}
        )
        
        json_response_backlog = backlog_response.json()

        if backlog_response.status_code == 200:
            created_issues.extend(json_response_backlog['issues'])
            if json_response_backlog['total'] == len(created_issues):
                break
            else:
                startAt_backlog += maxResults_backlog
        else:
            break

    startAt_backlog = 0
    while True:
        jql_query_backlog = f"project = '10000' AND resolutiondate >= '{current_day_start}' AND resolutiondate < '{next_day_start}'"
        jql_query_encoded_backlog = urllib.parse.quote(jql_query_backlog)
        api_url_backlog = f"{jira_url_backlog}{api_endpoint_backlog}?jql={jql_query_encoded_backlog}&startAt={startAt_backlog}&maxResults={maxResults_backlog}"

        backlog_response = requests.get(
            api_url_backlog,
            auth=HTTPBasicAuth(jira_email, jira_key),
            headers={"Accept": "application/json"}
        )
        json_response_backlog = backlog_response.json()
        print(f"Status code: {backlog_response.status_code}")

        if backlog_response.status_code == 200:
            resolved_issues.extend(json_response_backlog['issues'])
            if json_response_backlog['total'] == len(resolved_issues):
                break
            else:
                startAt_backlog += maxResults_backlog
        else:
            break

    current_started = previous_started
    created_count = len(created_issues)
    resolved_count = len(resolved_issues)
    total_for_day = current_started + created_count - resolved_count
    previous_started = total_for_day

    backlog_df = pd.DataFrame({
        'Day': [date],
        'Started': [current_started],
        'Created': [created_count],
        'Resolved': [resolved_count],
        'Total': [total_for_day]
    })

    backlog_df_temp.append(backlog_df)

final_backlog_df = pd.concat(backlog_df_temp).reset_index(drop=True)

today_backlog = datetime.now().date()
final_backlog_df['Day'] = pd.to_datetime(final_backlog_df['Day'], format='%d-%b-%Y').dt.date
filtered_backlog_df = final_backlog_df[final_backlog_df['Day'] <= today_backlog]

# Assuming filtered_backlog_df has a 'Total' column
totals = filtered_backlog_df['Total'].tolist()

# Determine the size of the figure based on the number of rows and columns
totals = filtered_backlog_df['Total'].tolist()

# Determine the size of the figure based on the number of rows and columns
num_rows, num_cols = filtered_backlog_df.shape

cell_width = 1.0
cell_height = 1.2

figwidth = num_cols * cell_width
figheight = num_rows * cell_height + 1

fig = plt.figure(figsize=(figwidth, figheight))

# Dynamic height ratios - more space for the chart and less for the table
ratio_for_chart = .7  # You can adjust this value
ratio_for_table = 1 - ratio_for_chart
gs = gridspec.GridSpec(2, 1, height_ratios=[num_rows*ratio_for_chart, num_rows*ratio_for_table])

ax1 = fig.add_subplot(gs[0])

# Adjust tick spacing based on data size
tick_spacing = max(1, len(totals) // 10)
ax1.xaxis.set_major_locator(plt.MaxNLocator(nbins=tick_spacing))

ax1.plot(totals, '-o', color='blue')
ax1.set_title("Total Issues Over Time", fontsize=10)
ax1.set_xticks(range(len(totals)))
ax1.set_xticklabels(filtered_backlog_df['Day'].tolist(), rotation=45, fontsize=8)
ax1.tick_params(axis='y', labelsize=8)

ax2 = fig.add_subplot(gs[1])
ax2.axis('tight')
ax2.axis('off')

the_table = ax2.table(cellText=filtered_backlog_df.values,
                      colLabels=filtered_backlog_df.columns,
                      loc='center',
                      cellLoc='center',
                      colWidths=[cell_width/num_cols]*num_cols)

ax2.text(0.5, 1.05, "Aging Backlog", size=10, ha="center", transform=ax2.transAxes)

# Adjust spacing
plt.subplots_adjust(hspace=0.5)

plt.tight_layout()
plt.savefig('aging_backlog.png', dpi=300, bbox_inches='tight')

#Extract what is needed from JIRA payload - store in dataframe
if isinstance(json_response, str):
    json_response = json.loads(json_response)

issues = all_issues

if isinstance(issues, list):
    data = []

    for issue in issues:
        key = issue['key']
        customfield_10035_obj = issue['fields'].get('customfield_10035', None)
        customfield_10035 = customfield_10035_obj['value'] if customfield_10035_obj else None
        created = issue['fields'].get('created', None)
        updated = issue['fields'].get('updated', None)
        resolved = issue['fields'].get('resolutiondate',None)
        assignee_obj = issue['fields'].get('assignee', None)
        assignee_name = assignee_obj['displayName'] if assignee_obj and 'displayName' in assignee_obj else None
        customfield_10300 = issue['fields'].get('customfield_10300', None)

        status_snapshot = issue['fields'].get('status', {}).get('name', None)

        org_obj = issue['fields'].get('customfield_10002', None)
        organization_name = org_obj[0]['name'] if org_obj else None

        component_obj = issue['fields'].get('components', None)
        component_name = component_obj[0]['name'] if component_obj else None

        data.append([key, customfield_10035, created, updated, resolved, assignee_name, customfield_10300, status_snapshot, organization_name, component_name])

    jira_df = pd.DataFrame(data, columns=['key', 'customfield_10035', 'created', 'updated', 'resolved', 'assignee', 'customfield_10300', 'status_snapshot', 'organization', 'component'])

#Determine tickets created and closed, merge by assignee, do some calculations and formatting
jira_df['TTR'] = jira_df.apply(lambda row: 
                               (pd.to_datetime(row['resolved']) - pd.to_datetime(row['created'])).total_seconds() / 3600
                               if row['resolved'] is not None else None, axis=1)
tickets_created = jira_df.groupby('assignee')['key'].count().reset_index(name='Tickets Created')
tickets_closed = jira_df[jira_df['status_snapshot'] == 'Resolved'].groupby('assignee')['key'].count().reset_index(name='Tickets Closed')
average_ttr = (jira_df[(jira_df['status_snapshot'] == 'Resolved') & (jira_df['TTR'].notnull())]
               .groupby('assignee')['TTR']
               .mean()
               .round(2)
               .reset_index(name='Average TTR (Hours)'))
user_grouped_data = pd.merge(tickets_created, tickets_closed, on='assignee', how='left')
user_grouped_data = pd.merge(user_grouped_data, average_ttr,on='assignee', how='left')
user_grouped_data['Tickets Closed'].fillna(0, inplace=True)
user_grouped_data['Tickets Closed'] = user_grouped_data['Tickets Closed'].astype(int)
user_grouped_data['Close to Created Ratio'] = user_grouped_data['Tickets Closed']/user_grouped_data['Tickets Created']
user_grouped_data['Close to Created Ratio'].replace([float('inf'), float('-inf')], 0, inplace=True)  # Replace infinite values with 0
user_grouped_data['Close to Created Ratio'] = (user_grouped_data['Close to Created Ratio'] * 100).round(2)
user_grouped_data = user_grouped_data.sort_values(by='Close to Created Ratio', ascending=False)
user_grouped_data['Close to Created Ratio'] = user_grouped_data['Close to Created Ratio'].astype(str) + '%'

#Function to color code Close to Created Ratio - Create second plot - Tickets Closed/Created by Agent 
def get_color(value):
    value = float(value.strip('%'))
    if value >= 90:
        return 'green'
    elif 75 <= value < 90:
        return (1.0, 0.76, 0.03)
    else:
        return 'red'

fig, ax = plt.subplots(figsize=(7, 5))
ax.axis('tight')
ax.axis('off')
the_table = ax.table(cellText=user_grouped_data.values, colLabels=user_grouped_data.columns, loc='center')

for i, value in enumerate(user_grouped_data["Close to Created Ratio"].values):
    color = get_color(value)
    the_table[(i + 1, user_grouped_data.columns.get_loc("Close to Created Ratio"))].set_facecolor(color)

plt.savefig('user_grouped_data.png', dpi=300, bbox_inches='tight')

#Create new dataframe with only Unresolved tickets, aggregate and group by channel
unresolved = jira_df[jira_df['status_snapshot'] != 'Resolved']
unresolved_by_channel = unresolved.groupby('customfield_10035')['key'].count().reset_index()
unresolved_by_channel.columns = ['Channel', 'Unresolved Count']

unresolved = unresolved.rename(columns={
    'customfield_10035': 'status',
    'customfield_10300': 't1_transfer_date'
})

#Create third plot image - Unresolved tickets by channel
plt.figure(figsize=(10, 5))
unresolved_by_channel = unresolved_by_channel.sort_values(by='Unresolved Count', ascending=False)
sns.heatmap(unresolved_by_channel.set_index('Channel'), annot=True, cmap='Blues', cbar=False, fmt='d')
plt.xticks([])
plt.title('Unresolved Tickets by Channel')
plt.ylabel(' ')
plt.tight_layout()
plt.savefig('unresolved_by_channel.png', pad_inches=0.05)

#Pivot AWS contacts queued, aggregate and group by hour of the day - store in dataframe
pivoted_result = pd.DataFrame(result)
pivoted_result['Timestamp'] = pd.to_datetime(pivoted_result['Timestamp'])
pivoted_result['Hour'] = pivoted_result['Timestamp'].dt.hour
hourly_sum = pivoted_result.groupby('Hour')['Contacts Queued'].sum()
default_df = pd.DataFrame(np.zeros((1, 24)), columns=range(24), index=['Phone'])
for hour, value in hourly_sum.items():
    default_df[hour] = value

#Pivot JIRA ticket data less Phone, aggregate and group by type and hour of the day - store in dataframe
jira_df['created'] = pd.to_datetime(jira_df['created'], infer_datetime_format=True)
jira_df['hour'] = jira_df['created'].dt.hour
df_filtered = jira_df[jira_df['customfield_10035'].isin(['Email', 'Portal','Live Chat'])]
grouped = df_filtered.groupby(['customfield_10035', 'hour']).size().reset_index(name='count')
pivot_table = grouped.pivot_table(index='customfield_10035', columns='hour', values='count', fill_value=0)
pivot_table.columns = [str(col) for col in pivot_table.columns]
pivot_table.columns = pivot_table.columns.astype(int)
pivot_table = pivot_table.reindex(columns=range(24), fill_value=0)
jira_df = jira_df.reindex(columns=range(24), fill_value=0)
final_df = pd.concat([pivot_table, default_df])
final_df.columns = [f'{hour%12 if hour%12 != 0 else 12}{("am" if hour < 12 else "pm")}' for hour in final_df.columns]
final_df = final_df.astype(int)

#Do some cleanup to the hour buckets
final_df['12am-6am'] = final_df[['12am', '1am', '2am', '3am', '4am', '5am', '6am']].sum(axis=1)
final_df['7pm-11pm'] = final_df[['7pm', '8pm', '9pm', '10pm', '11pm']].sum(axis=1)
final_df = final_df.drop(columns=['12am', '1am', '2am', '3am', '4am', '5am', '6am', '7pm', '8pm', '9pm', '10pm', '11pm'])
final_df = final_df[['12am-6am'] + [col for col in final_df if col not in ['12am-6am', '7pm-11pm']] + ['7pm-11pm']]

#Create fourth plot image - Ticket vol by Hour and Channel
fig, axes = plt.subplots(final_df.shape[0], 1, figsize=(10, 8))

for index, (row_name, row_data) in enumerate(final_df.iterrows()):
    ax = axes[index]
    sns.heatmap(row_data.to_frame().T, annot=True, fmt="d", cmap="coolwarm", cbar=False, ax=ax)
    ax.set_title(f'{row_name}')
    ax.set_yticklabels([])

    if index < final_df.shape[0] - 1:
        ax.set_xticklabels([])
    else:
        ax.set_xticklabels(ax.get_xticklabels(), rotation=45)

plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.savefig('combined_heatmaps_without_title.png')
plt.close()
img = Image.open('combined_heatmaps_without_title.png')
draw = ImageDraw.Draw(img)
font = ImageFont.truetype("arial.ttf", 16)
text = "Vol by Hour by Channel"
textwidth, textheight = draw.textsize(text, font)
x = (img.width - textwidth) / 2
y = 2
draw.text((x, y), text, font=font, fill="black")

img.save('combined_heatmaps.png')

#Get Yesterday for email formatting
get_yesterday = datetime.now() - relativedelta(days=1)
yesterday = format(get_yesterday, '%D')

get_today = datetime.now()
today = format(get_today, '%D')

#Begin Email Server Development
smtp_port = 587
smtp_server = "smtp.gmail.com"

with open('C:\\Users\\\mdunlap\Desktop\\Keys\\Support Daily Readout\\hha email - from.txt', 'r') as from_email2:
    from_email = from_email2.read().replace('\n', '')

email_from = from_email

#List for easy addition of members
with open('C:\\Users\\\mdunlap\Desktop\\Keys\\Support Daily Readout\\hha email - to.txt', 'r') as to_email_init:
    to_email = to_email_init.read().strip().split(',')

email_members = to_email

with open('C:\\Users\\\mdunlap\Desktop\\Keys\\Support Daily Readout\\hha email - to_2.txt', 'r') as to_email_project:
    to_email_2 = to_email_project.read().strip().split(',')
    
email_members_2 = to_email_2

with open('C:\\Users\\\mdunlap\Desktop\\Keys\\Support Daily Readout\\hha email - to_elt.txt', 'r') as to_email_elt:
    to_email_3 = to_email_elt.read().strip().split(',')
    
email_members_3 = to_email_3

with open('C:\\Users\\\mdunlap\Desktop\\Keys\\Support Daily Readout\\google app password.txt', 'r') as app_pass:
    google_pass = app_pass.read().replace('\n', '')

google_password = google_pass

subject = 'Support Daily Readout' + ' '+ '-' + ' ' + yesterday

report_name = 'Support Daily Readout'

#Send email functions, attach images created above and send to recipient list
def create_xlsx_with_images():
    wb = openpyxl.Workbook()
    ws = wb.active

    images = ["unresolved_by_channel.png", "combined_heatmaps.png", "contacts_queued.png","user_grouped_data.png"]

    for idx, img_path in enumerate(images, 1):
        img = XLSXImage(img_path)
        ws.column_dimensions[get_column_letter(idx)].width = 12
        ws.row_dimensions[idx].height = img.height
        ws.add_image(img, f'{get_column_letter(idx)}{idx}')

    wb.save(report_name+".xlsx")

jira_details_filename = f"Unresolved Details.xlsx"
unresolved.to_excel(jira_details_filename)

def send_dailyreadout(email_members):
    msg = MIMEMultipart()
    msg['From'] = email_from
    msg['To'] = ', '.join(email_members)
    msg['Subject'] = subject
        
    body = 'Please find the daily Support metrics readout for' + ' ' + today +'.'
    msg.attach(MIMEText(body, 'plain'))
        
    # Attach the XLSX to the email
    create_xlsx_with_images()
        
    for filename in [report_name + ".xlsx", jira_details_filename]: 
            with open(filename, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename= {filename}",
                )

                msg.attach(part)
            
    text = msg.as_string()

    mail_server = smtplib.SMTP(smtp_server, smtp_port)
    mail_server.starttls()
    mail_server.login(email_from, google_password)
    mail_server.sendmail(email_from, email_members, text)
    mail_server.quit()


def send_aging_backlog(email_members_2):
    subject = 'Aging Backlog Status Update' + ' ' + '-' +' ' + today

    body = 'Aging Backlog Status Update as of' + ' ' + today + '.'
    
    msg = MIMEMultipart()
    msg['From'] = email_from
    msg['To'] = ', '.join(email_members_2)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    with open("aging_backlog.png", "rb") as attachment:
        part = MIMEImage(attachment.read())
        part.add_header("Content-Disposition", f"attachment; filename= aging_backlog.png")
        msg.attach(part)

    text = msg.as_string()

    mail_server = smtplib.SMTP(smtp_server, smtp_port)
    mail_server.starttls()
    mail_server.login(email_from, google_password)
    mail_server.sendmail(email_from, email_members_2, text)
    mail_server.quit()

def send_aging_backlog_elt(email_members_3):
    subject = 'ELT - Aging Backlog Status Update' + ' ' + '-' +' ' + today

    body = 'Aging Backlog Status Update as of' + ' ' + today + '.'
    
    msg = MIMEMultipart()
    msg['From'] = email_from
    msg['To'] = ', '.join(email_members_3)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    with open("aging_backlog.png", "rb") as attachment:
        part = MIMEImage(attachment.read())
        part.add_header("Content-Disposition", f"attachment; filename= aging_backlog.png")
        msg.attach(part)

    text = msg.as_string()

    mail_server = smtplib.SMTP(smtp_server, smtp_port)
    mail_server.starttls()
    mail_server.login(email_from, google_password)
    mail_server.sendmail(email_from, email_members_3, text)
    mail_server.quit()

send_dailyreadout(email_members)

send_aging_backlog(email_members_2)

send_aging_backlog_elt(email_members_3)
